from student_rewards_classes import Student, Reward, Task
from table_creator import create_student_example, create_task_example, create_reward_example, create_table_dict
import openpyxl


def print_worksheet(workbook):
    """
    this function prints the active worksheet of the workbook
    :param workbook: the .xlsx file address
    """

    wb = openpyxl.load_workbook(workbook)
    ws = wb.active

    for row in ws.iter_rows(values_only=True):
        for cell in row:
            print(cell, end=" ")
        print()


def insert_completed_task(worksheet, student_name, task_name, value):
    """
    Updates a completed task to the .xlsx file at the right row and column. It uses strings as indexes
    :param worksheet: the .xlsx file address
    :param student_name: the student.name attribute
    :param task_name: the task.name attribute
    :param value: the completed value
    :return: a new .xlsx file
    """
    wb = openpyxl.load_workbook(worksheet)
    ws = wb.active

    row_label = student_name
    column_label = task_name
    value_to_insert = value

    column_index = None
    for cell in ws[1]:  # Assuming the first row contains column labels
        if cell.value == column_label:
            column_index = cell.column
            break

    if column_index:
        for rowIndex, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row[0] == row_label:
                # rowIndex gives the correct row number since enumeration starts at 1
                ws.cell(row=rowIndex, column=column_index, value=value_to_insert)
                break

    wb.save("updated_student_reward_table.xlsx")


def write_rows():
    """
    this function creates the table with all the students and task. It iterates through the get_all_* methods
    building rows for students names and columns for tasks names.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    for task in Task.get_all_tasks():
        print(task.name)

    for col, task in enumerate(Task.get_all_tasks(), start=2):
        ws.cell(row=1, column=col, value=task.name)

    for row, student in enumerate(Student.get_all_students(), start=2):
        ws.cell(row=row, column=1, value=student.name)

    wb.save("student_reward_table.xlsx")
    return 0


def main():
    """
    this function creates a bunch of random students and tasks
    """
    task_generator = create_task_example()
    student_generator = create_student_example()
    for i in range(10):
        student = next(student_generator)
        task = next(task_generator)
        task.set_requirements(i)

    for student in Student.get_all_students():
        student.update_tasks()

    return 0


if __name__ == "__main__":
    main()
    write_rows()
    #print_worksheet("student_reward_table.xlsx")
    worksheet = "student_reward_table.xlsx"
    student_name = 'Olivia'
    task_name = 'Burpees'
    value = 'Completed'
    insert_completed_task("student_reward_table.xlsx", student_name, task_name, value)

    student_list = [student for student in Student.get_all_students()]
    for student in student_list:
        student_name = student.name
        print(student_name, end=' ; ')
        insert_completed_task("updated_student_reward_table.xlsx", student_name, 'Squats', value)

    print_worksheet("updated_student_reward_table.xlsx")